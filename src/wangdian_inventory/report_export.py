"""Excel exports for the daily and weekly supply-chain reports."""

import math
from datetime import date, timedelta
from io import BytesIO
from typing import Any, Callable, Dict, Iterable, List, Mapping, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


NAVY = "17365D"
BLUE = "1F4E78"
LIGHT_BLUE = "D9EAF7"
GREEN = "548235"
LIGHT_GREEN = "E2F0D9"
GOLD = "BF9000"
LIGHT_GOLD = "FFF2CC"
RED = "C00000"
LIGHT_RED = "FCE4D6"
ORANGE = "ED7D31"
LIGHT_ORANGE = "F4B183"
PURPLE = "7030A0"
LIGHT_PURPLE = "E4DFEC"
GRAY = "666666"
LIGHT_GRAY = "F2F2F2"
WHITE = "FFFFFF"

HEADER_FILL = PatternFill("solid", fgColor=BLUE)
HEADER_FONT = Font(color=WHITE, bold=True)
THIN_BORDER = Border(bottom=Side(style="thin", color="D9E2F3"))


def _value(value: Any) -> Any:
    return "" if value is None else value


def _number(value: Any) -> float:
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def _integer(value: Any) -> Any:
    """Return a spreadsheet-safe integer while preserving empty/text values."""
    if value in (None, ""):
        return value
    try:
        number = float(value)
    except (TypeError, ValueError):
        return value
    # Python's round uses bankers rounding; report quantities use half-up.
    return int(math.floor(number + 0.5)) if number >= 0 else int(math.ceil(number - 0.5))


def _percent(numerator: Any, denominator: Any) -> float | None:
    total = _number(denominator)
    return _number(numerator) / total if total > 0 else None


def _rounded_days(value: Any) -> int | None:
    """Round estimated coverage days to a conventional whole-day value."""
    if value in (None, ""):
        return None
    number = _number(value)
    return int(math.floor(number + 0.5)) if number >= 0 else int(math.ceil(number - 0.5))


def _spec_name(row: Mapping[str, Any]) -> str:
    """Display missing product variants explicitly in SKU-level report rows."""
    return str(row.get("spec_name") or "未维护")


def _spec_remark(row: Mapping[str, Any]) -> str:
    """Display the SKU classification maintained in WangDian's remark field."""
    return str(row.get("spec_remark") or "未维护")


REPORT_SALES_WAREHOUSES = {
    "主仓库（新）",
    "TMT康复仓",
    "杰菲克德国仓",
    "日本仓",
    "ENYISA仓",
}


def _format_cell(cell: Any) -> None:
    if isinstance(cell.value, (int, float)):
        cell.number_format = "#,##0"
        cell.alignment = Alignment(horizontal="right", vertical="center")
    else:
        cell.alignment = Alignment(vertical="center")


def _fit_columns(sheet: Any, max_width: int = 30) -> None:
    # A report may include tens of thousands of inventory rows. A representative
    # prefix is enough for readable widths and avoids making export time grow with
    # every historical row.
    for index, column in enumerate(sheet.iter_cols(max_row=min(sheet.max_row, 300)), 1):
        width = max((len(str(cell.value or "")) for cell in column), default=8) + 2
        sheet.column_dimensions[get_column_letter(index)].width = min(max(width, 10), max_width)


def _write_table(
    workbook: Workbook,
    name: str,
    headers: Sequence[str],
    rows: Iterable[Sequence[Any]],
    *,
    row_fill: Callable[[Sequence[Any]], str | None] | None = None,
    percent_headers: Sequence[str] = (),
    integer_headers: Sequence[str] = (),
    decimal_headers: Sequence[str] = (),
    positive_fill_headers: Sequence[str] = (),
) -> Any:
    sheet = workbook.create_sheet(name[:31])
    sheet.sheet_view.showGridLines = False
    sheet.append(list(headers))
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    percent_indexes = {headers.index(header) + 1 for header in percent_headers if header in headers}
    integer_indexes = {headers.index(header) + 1 for header in integer_headers if header in headers}
    decimal_indexes = {headers.index(header) + 1 for header in decimal_headers if header in headers}
    positive_fill_indexes = {headers.index(header) + 1 for header in positive_fill_headers if header in headers}
    row_index = 2
    for row in rows:
        values = [_value(value) for value in row]
        for index in integer_indexes:
            values[index - 1] = _integer(values[index - 1])
        sheet.append(values)
        fill_color = row_fill(row) if row_fill else None
        for index in range(1, len(headers) + 1):
            cell = sheet.cell(row_index, index)
            _format_cell(cell)
            cell.border = THIN_BORDER
            if index in percent_indexes and isinstance(cell.value, (int, float)):
                cell.number_format = "0.0%"
            elif index in integer_indexes and isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0"
            elif index in decimal_indexes and isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0.###"
            if index in positive_fill_indexes and isinstance(cell.value, (int, float)) and cell.value > 0:
                cell.fill = PatternFill("solid", fgColor=LIGHT_GOLD)
            if fill_color:
                cell.fill = PatternFill("solid", fgColor=fill_color)
        row_index += 1
    sheet.freeze_panes = "A2"
    if sheet.max_row >= 2:
        sheet.auto_filter.ref = sheet.dimensions
    _fit_columns(sheet)
    sheet.row_dimensions[1].height = 24
    return sheet


def _product_rows(database: Any) -> List[Dict[str, Any]]:
    with database.connect() as connection:
        rows = connection.execute(
            """
            SELECT i.warehouse_id, i.warehouse_no, i.warehouse_name, i.sku_no,
                   p.goods_no, p.short_name, p.goods_name, p.spec_name, p.category,
                   p.product_structure, p.spec_remark, p.supplier_name, p.purchase_price,
                   p.erp_price, p.retail_price,
                   i.stock_num, i.available_num, i.purchase_in_transit_num,
                   COALESCE(NULLIF(i.avg_cost_price,0),i.cost_price,0) erp_cost_price,
                   i.stock_num * COALESCE(NULLIF(i.avg_cost_price,0),i.cost_price,0) erp_stock_value,
                   i.synced_at
            FROM inventory_current i
            JOIN products p ON p.sku_no=i.sku_no
            LEFT JOIN warehouse_master wm ON wm.warehouse_id=i.warehouse_id
            WHERE COALESCE(wm.is_disabled,0)=0 AND COALESCE(wm.role,'sales')='sales'
            ORDER BY i.warehouse_name,p.short_name,i.sku_no
            """
        ).fetchall()
    return [dict(row) for row in rows]


def _structure_rows(
    analysis_rows: Sequence[Mapping[str, Any]], product_by_sku: Mapping[str, Mapping[str, Any]]
) -> List[Dict[str, Any]]:
    """Produce operating candidates without mixing in purchase quantities or MOQ."""
    candidates: List[Dict[str, Any]] = []
    for source in analysis_rows:
        row = dict(source)
        product = product_by_sku.get(row.get("sku_no"), {})
        structure = row.get("product_structure") or product.get("product_structure") or "待补充"
        if structure == "清仓款":
            continue
        sales_30 = _number(row.get("sales_30d_qty"))
        sales_7 = _number(row.get("sales_7d_qty") or row.get("sales_qty"))
        returns = _number(row.get("return_30d_qty") or row.get("return_qty"))
        trend = row.get("trend_coefficient")
        trend_value = _number(trend) if trend is not None else None
        coverage = row.get("projected_coverage_days")
        coverage_value = _number(coverage) if coverage is not None else None
        retail = _number(product.get("retail_price"))
        purchase_price = _number(product.get("purchase_price"))
        gross_profit = retail - purchase_price if purchase_price > 0 else None
        gross_margin = gross_profit / retail if gross_profit is not None and retail > 0 else None
        return_rate = _percent(returns, sales_30)
        candidate = ""
        reasons: List[str] = []
        if _number(row.get("available_num")) > 0 and (sales_30 == 0 or (coverage_value is not None and coverage_value > 90)):
            candidate = "清仓候选"
            if sales_30 == 0:
                reasons.append("近30日无发货销量")
            if coverage_value is not None and coverage_value > 90:
                reasons.append(f"含在途库存可覆盖 {coverage_value:.0f} 天")
            if return_rate is not None and return_rate > 0.10:
                reasons.append(f"退货率偏高 {return_rate:.1%}")
            if gross_margin is not None and gross_margin <= 0:
                reasons.append("采购价不低于零售价")
        elif (
            structure != "放大款" and sales_30 >= 30 and trend_value is not None and trend_value >= 1.2
            and (return_rate is None or return_rate <= 0.10) and (gross_profit is None or gross_profit > 0)
        ):
            candidate = "放大款候选"
            reasons.append(f"近7日趋势系数 {trend_value:.2f}")
            reasons.append(f"近30日销量 {sales_30:.0f}")
            if return_rate is not None:
                reasons.append(f"退货率 {return_rate:.1%}")
        if candidate:
            candidates.append({
                "candidate": candidate,
                "structure": structure,
                "warehouse_name": row.get("warehouse_name"),
                "sku_no": row.get("sku_no"),
                "short_name": row.get("short_name") or product.get("short_name"),
                "spec_name": row.get("spec_name") or product.get("spec_name"),
                "supplier_name": row.get("supplier_name") or product.get("supplier_name"),
                "sales_30": sales_30,
                "sales_7": sales_7,
                "trend": trend_value,
                "return_qty": returns,
                "return_rate": return_rate,
                "retail_price": retail or None,
                "purchase_price": purchase_price or None,
                "gross_profit": gross_profit,
                "gross_margin": gross_margin,
                "available_num": _number(row.get("available_num")),
                "coverage": coverage_value,
                "reason": "；".join(reasons),
            })
    candidates.sort(key=lambda row: (0 if row["candidate"] == "清仓候选" else 1, -row["sales_30"], row["sku_no"] or ""))
    return candidates


def _report_data(database: Any, start_date: str, end_date: str) -> Dict[str, Any]:
    dashboard = database.dashboard(start_date, end_date, limit=2000)
    sales = database.warehouse_sku_sales(start_date, end_date, limit=100000)
    sales_day_count = (date.fromisoformat(end_date) - date.fromisoformat(start_date)).days + 1
    sales_items = []
    for row in sales["items"]:
        if row.get("warehouse_name") not in REPORT_SALES_WAREHOUSES:
            continue
        item = dict(row)
        # Report 04 uses a deliberately narrow demand baseline: prefer the
        # last 7 calendar days, then fall back to the last 3 days, then to the
        # actual number of days with historical sales. This keeps the warning
        # tied to recent warehouse movement instead of the selected range's
        # full-period average.
        sales_7 = _number(item.get("sales_7d_qty"))
        sales_3 = _number(item.get("sales_3d_qty"))
        history_days = int(_number(item.get("sales_history_days")))
        if sales_7 > 0:
            daily_sales = sales_7 / 7
        elif history_days >= 3:
            daily_sales = sales_3 / 3
        elif history_days > 0:
            daily_sales = _number(item.get("sales_qty")) / history_days
        else:
            daily_sales = 0.0
        item["daily_sales"] = daily_sales
        item["inventory_days_reason"] = "日销量为零" if daily_sales <= 0 else ""
        item["inventory_remaining_days"] = (
            _number(item.get("available_num")) / daily_sales if daily_sales > 0 else None
        )
        item["inventory_with_transit_days"] = (
            (_number(item.get("available_num")) + _number(item.get("purchase_in_transit_num"))) / daily_sales
            if daily_sales > 0 else None
        )
        sales_items.append(item)
    sales = dict(sales)
    sales["items"] = sales_items
    inbound = database.inbound_analysis(start_date, end_date, limit=100000)
    clearance = database.replenishment_analysis(
        start_date, end_date, mode="replenishment", alert_mode="clearance", limit=100000
    )
    normal_alerts = database.replenishment_analysis(
        start_date, end_date, mode="replenishment", alert_mode="normal", limit=100000
    )
    planning_rows = database._warehouse_planning_rows(start_date, end_date)["rows"]
    return_start = (date.fromisoformat(end_date) - timedelta(days=29)).isoformat()
    with database.connect() as connection:
        return_rows = connection.execute(
            """
            SELECT m.sku_no, m.warehouse_id, SUM(m.in_num) return_qty
            FROM movements m
            LEFT JOIN warehouse_master wm ON wm.warehouse_id=m.warehouse_id
            WHERE m.movement_date BETWEEN ? AND ? AND m.movement_type=3
              AND COALESCE(wm.is_disabled,0)=0 AND COALESCE(wm.role,'sales')='sales'
            GROUP BY m.sku_no,m.warehouse_id
            """,
            (return_start, end_date),
        ).fetchall()
    return_by_key = {(row["sku_no"], row["warehouse_id"]): _number(row["return_qty"]) for row in return_rows}
    for row in planning_rows:
        row["return_30d_qty"] = return_by_key.get((row["sku_no"], row["warehouse_id"]), 0.0)
    purchase = database.purchase_plan(start_date, end_date, limit=100000)
    inventory = _product_rows(database)
    product_by_sku: Dict[str, Dict[str, Any]] = {}
    for row in inventory:
        product_by_sku.setdefault(row["sku_no"], row)
    return {
        "dashboard": dashboard,
        "sales": sales,
        "inbound": inbound,
        "clearance": clearance,
        "normal_alerts": normal_alerts,
        "planning_rows": planning_rows,
        "purchase": purchase,
        "inventory": inventory,
        "structure_rows": _structure_rows(planning_rows, product_by_sku),
    }


def _summary_sheet(
    workbook: Workbook, report_label: str, start_date: str, end_date: str, data: Mapping[str, Any]
) -> None:
    sheet = workbook.create_sheet(f"01_{report_label}总览")
    sheet.sheet_view.showGridLines = False
    sheet.merge_cells("A1:L1")
    sheet["A1"] = f"旺店通供应链{report_label}"
    sheet["A1"].fill = PatternFill("solid", fgColor=NAVY)
    sheet["A1"].font = Font(size=20, bold=True, color=WHITE)
    sheet["A1"].alignment = Alignment(horizontal="left", vertical="center")
    sheet.row_dimensions[1].height = 34
    sheet.merge_cells("A2:L2")
    sheet["A2"] = f"统计范围：{start_date} 至 {end_date}    |    库存为最新销售仓快照，销量按实际发货日，退货按退货入库日"
    sheet["A2"].font = Font(color=GRAY, italic=True)
    sheet["A2"].alignment = Alignment(vertical="center")

    summary = data["dashboard"]["summary"]
    purchase = data["purchase"]["summary"]
    return_rate = _percent(summary.get("movement_return_qty"), summary.get("movement_sales_qty"))
    card_groups = [
        ("📦 库存", BLUE, [
            ("库存 SKU", summary.get("sku_count", 0)),
            ("当前库存", summary.get("stock_num", 0)),
            ("可用库存", summary.get("available_num", 0)),
            ("采购在途", summary.get("purchase_in_transit_num", 0)),
        ]),
        ("📈 经营", GREEN, [
            ("发货销量", summary.get("movement_sales_qty", 0)),
            ("退货入库", summary.get("movement_return_qty", 0)),
            ("净销量", _number(summary.get("movement_sales_qty")) - _number(summary.get("movement_return_qty"))),
            ("退货率", return_rate),
        ]),
        ("⚠️ 风险", RED, [
            ("库存积压预警", len(data["normal_alerts"]["items"])),
            ("清仓预警", len(data["clearance"]["items"])),
            ("严重缺货", purchase.get("severe_shortage_count", 0)),
            ("结构调整候选", len(data["structure_rows"])),
        ]),
        ("🛒 采购", GOLD, [
            ("计算采购量", purchase.get("total_calculated_order_qty", 0)),
            ("起订量补足", purchase.get("total_moq_uplift_qty", 0)),
            ("建议下单量", purchase.get("total_order_qty", 0)),
            ("采购入库", summary.get("purchase_qty", 0)),
        ]),
    ]
    column_sets = [("A", "F", 4), ("G", "L", 4), ("A", "F", 9), ("G", "L", 9)]
    for (title, color, metrics), (start, end, top_row) in zip(card_groups, column_sets):
        sheet.merge_cells(f"{start}{top_row}:{end}{top_row}")
        header = sheet[f"{start}{top_row}"]
        header.value = title
        header.fill = PatternFill("solid", fgColor=color)
        header.font = Font(color=WHITE, bold=True, size=12)
        header.alignment = Alignment(horizontal="center")
        for offset, (metric, amount) in enumerate(metrics):
            row = top_row + 1 + offset
            value_column = end if start == "A" else "J"
            label_end = get_column_letter(ord(value_column) - 64 - 1)
            sheet.merge_cells(f"{start}{row}:{label_end}{row}")
            # The two left columns hold the label; the last column holds its value.
            sheet[f"{start}{row}"] = metric
            sheet[f"{start}{row}"].fill = PatternFill("solid", fgColor=LIGHT_GRAY)
            sheet[f"{start}{row}"].font = Font(color=GRAY)
            sheet[f"{start}{row}"].alignment = Alignment(vertical="center")
            value_cell = sheet[f"{value_column}{row}"]
            value_cell.value = amount
            value_cell.font = Font(size=12, bold=True, color=color)
            value_cell.alignment = Alignment(horizontal="right", vertical="center")
            value_cell.number_format = "0.0%" if metric == "退货率" and amount is not None else "#,##0"

    sheet.merge_cells("A15:F15")
    sheet["A15"] = "🏭 仓库库存摘要"
    sheet["A15"].fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    sheet["A15"].font = Font(bold=True, color=NAVY)
    sheet.merge_cells("G15:L15")
    sheet["G15"] = "🚨 优先处理事项"
    sheet["G15"].fill = PatternFill("solid", fgColor=LIGHT_RED)
    sheet["G15"].font = Font(bold=True, color=RED)
    warehouse_totals: Dict[str, Dict[str, float]] = {}
    for row in data["inventory"]:
        key = row.get("warehouse_name") or "仓库待补充"
        current = warehouse_totals.setdefault(key, {"sku": 0, "stock": 0, "available": 0, "transit": 0})
        current["sku"] += 1
        current["stock"] += _number(row.get("stock_num"))
        current["available"] += _number(row.get("available_num"))
        current["transit"] += _number(row.get("purchase_in_transit_num"))
    for column, value in enumerate(["仓库", "SKU数", "当前库存", "可用库存", "采购在途", "", "优先级", "事项", "数量", "", "", ""], 1):
        sheet.cell(16, column, value)
    for cell in sheet[16]:
        if cell.column <= 5 or 7 <= cell.column <= 9:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")
    priorities = [
        ("高", "交期内预计缺货", purchase.get("severe_shortage_count", 0), LIGHT_RED),
        ("中", "本周需要下单", purchase.get("due_within_week_count", 0), LIGHT_ORANGE),
        ("中", "清仓预警", len(data["clearance"]["items"]), LIGHT_PURPLE),
        ("低", "库存积压预警", len(data["normal_alerts"]["items"]), LIGHT_GOLD),
    ]
    max_rows = max(len(warehouse_totals), len(priorities))
    for index in range(max_rows):
        row_index = 17 + index
        if index < len(warehouse_totals):
            name, amounts = sorted(warehouse_totals.items(), key=lambda item: -item[1]["available"])[index]
            values = [name, amounts["sku"], amounts["stock"], amounts["available"], amounts["transit"]]
            for col, value in enumerate(values, 1):
                cell = sheet.cell(row_index, col, value)
                _format_cell(cell)
                if col > 1:
                    cell.value = _integer(cell.value)
                    cell.number_format = "#,##0"
                cell.border = THIN_BORDER
        if index < len(priorities):
            priority, item, count, fill = priorities[index]
            for col, value in ((7, priority), (8, item), (9, count)):
                cell = sheet.cell(row_index, col, value)
                cell.fill = PatternFill("solid", fgColor=fill)
                cell.border = THIN_BORDER
                _format_cell(cell)
                if col == 9:
                    cell.value = _integer(cell.value)
                    cell.number_format = "#,##0"
    for column, width in {"A": 18, "B": 8, "C": 15, "D": 15, "E": 15, "F": 13, "G": 8, "H": 14, "I": 10, "J": 8, "K": 10, "L": 13}.items():
        sheet.column_dimensions[column].width = width
    sheet.freeze_panes = "A4"


def _build(database: Any, start_date: str, end_date: str, report_type: str) -> bytes:
    data = _report_data(database, start_date, end_date)
    workbook = Workbook()
    workbook.remove(workbook.active)
    report_label = "日报" if report_type == "daily" else "周报"
    _summary_sheet(workbook, report_label, start_date, end_date, data)

    _write_table(workbook, "02_运营_清仓预警", ["清仓状态", "仓库", "SKU", "商品简称", "规格", "规格备注", "供应商", "可用库存", "预计剩余天数", "趋势系数", "处理建议"], [
        [r.get("recommendation"), r.get("warehouse_name"), r.get("sku_no"), r.get("short_name"), _spec_name(r), _spec_remark(r), r.get("supplier_name"), r.get("available_num"), _rounded_days(r.get("forecast_coverage_days")), r.get("trend_coefficient"), "清仓款不参与采购，跟踪去化进度"]
        for r in data["clearance"]["items"]
    ], row_fill=lambda row: LIGHT_PURPLE, integer_headers=["可用库存", "预计剩余天数"], decimal_headers=["趋势系数"])

    alert_colors = {"严重积压": LIGHT_RED, "中度积压": LIGHT_ORANGE, "库存预警": LIGHT_GOLD}
    _write_table(workbook, "03_运营_库存积压预警", ["积压等级", "仓库", "SKU", "商品简称", "规格", "规格备注", "采购价", "ERP价", "供应商", "可用库存", "采购在途", "不含在途覆盖天数", "含在途覆盖天数", "趋势系数"], [
        [r.get("recommendation"), r.get("warehouse_name"), r.get("sku_no"), r.get("short_name"), _spec_name(r), _spec_remark(r), r.get("purchase_price"), r.get("erp_price"), r.get("supplier_name"), r.get("available_num"), r.get("purchase_in_transit_num"), _rounded_days(r.get("forecast_coverage_days")), _rounded_days(r.get("projected_coverage_days")), r.get("trend_coefficient")]
        for r in data["normal_alerts"]["items"]
    ], row_fill=lambda row: alert_colors.get(str(row[0]), LIGHT_GOLD), integer_headers=["可用库存", "采购在途", "不含在途覆盖天数", "含在途覆盖天数"], decimal_headers=["采购价", "ERP价", "趋势系数"])

    _write_table(workbook, "04_运营_采购在途预警", ["仓库", "SKU", "商品简称", "规格", "规格备注", "供应商", "日销量", "净销量", "退货率", "7日销量", "15日销量", "30日销量", "库存", "可用库存", "采购在途", "不含在途库存天数", "含在途库存天数"], [
        [r.get("warehouse_name"), r.get("sku_no"), r.get("short_name"), _spec_name(r), _spec_remark(r), r.get("supplier_name"), r.get("daily_sales"), r.get("net_sales_qty"), _percent(r.get("return_qty"), r.get("sales_qty")), r.get("sales_7d_qty"), r.get("sales_15d_qty"), r.get("sales_30d_qty"), r.get("stock_num"), r.get("available_num"), r.get("purchase_in_transit_num"), _rounded_days(r.get("inventory_remaining_days")) if r.get("inventory_remaining_days") is not None else r.get("inventory_days_reason"), _rounded_days(r.get("inventory_with_transit_days")) if r.get("inventory_with_transit_days") is not None else r.get("inventory_days_reason")]
        for r in data["sales"]["items"]
    ], percent_headers=["退货率"], integer_headers=["净销量", "7日销量", "15日销量", "30日销量", "库存", "可用库存", "采购在途", "不含在途库存天数", "含在途库存天数"], decimal_headers=["日销量"], positive_fill_headers=["采购在途"])

    _write_table(workbook, "05_运营_结构调整建议", ["建议结构", "现有结构", "仓库", "SKU", "商品简称", "规格", "供应商", "30日销量", "7日销量", "趋势系数", "退货量", "退货率", "零售价", "采购价", "单件毛利", "毛利率", "可用库存", "预计库存天数", "触发原因", "运营确认"], [
        [r["candidate"], r["structure"], r["warehouse_name"], r["sku_no"], r["short_name"], _spec_name(r), r["supplier_name"], r["sales_30"], r["sales_7"], r["trend"], r["return_qty"], r["return_rate"], r["retail_price"], r["purchase_price"], r["gross_profit"], r["gross_margin"], r["available_num"], _rounded_days(r["coverage"]), r["reason"], "待运营确认"]
        for r in data["structure_rows"]
    ], row_fill=lambda row: LIGHT_PURPLE if row[0] == "清仓候选" else LIGHT_GREEN, percent_headers=["退货率", "毛利率"], integer_headers=["30日销量", "7日销量", "退货量", "可用库存", "预计库存天数"], decimal_headers=["趋势系数", "零售价", "采购价", "单件毛利"])

    zero_sales_rows = sorted(
        (
            r for r in data["planning_rows"]
            if _number(r.get("sales_30d_qty")) == 0
            and (_number(r.get("available_num")) > 0 or _number(r.get("purchase_in_transit_num")) > 0)
        ),
        key=lambda r: (-(_number(r.get("available_num")) + _number(r.get("purchase_in_transit_num"))), r.get("sku_no") or ""),
    )[:2000]
    _write_table(workbook, "06_运营_零销量与新品", ["仓库", "SKU", "商品简称", "规格", "规格备注", "供应商", "现有结构", "库存", "采购在途", "剩余库存天数", "30日销量", "状态", "运营建议"], [
        [r.get("warehouse_name"), r.get("sku_no"), r.get("short_name"), _spec_name(r), _spec_remark(r), r.get("supplier_name"), r.get("product_structure"), r.get("stock_num"), r.get("purchase_in_transit_num"), _rounded_days(r.get("projected_coverage_days")) if r.get("projected_coverage_days") is not None else "无销量", r.get("sales_30d_qty"), "新品在途" if _number(r.get("stock_num")) == 0 and _number(r.get("purchase_in_transit_num")) > 0 else "零销量", "确认新品上架节奏或转入清仓候选"]
        for r in zero_sales_rows
    ], integer_headers=["库存", "采购在途", "剩余库存天数", "30日销量"])

    purchase_summary = data["purchase"]["summary"]
    purchase_colors = {
        "urgent": LIGHT_RED,
        "within_week": LIGHT_GOLD,
        "later": LIGHT_BLUE,
        "low_demand": LIGHT_PURPLE,
    }
    purchase_severity = {
        (r.get("warehouse_name"), r.get("sku_no")): r.get("procurement_severity")
        for r in data["purchase"]["items"]
    }
    purchase_sheet = _write_table(workbook, "07_采购_采购计划", ["优先级", "建议下单日", "目标仓", "SKU", "商品简称", "规格", "规格备注", "供应商", "产线", "产能", "生产天数", "可用库存", "采购在途", "建议调拨", "趋势系数", "日销量", "计算采购量", "最低起订量", "起订量补足", "建议下单量", "起订量处理", "剩余库存天数（含在途）", "预计缺货日", "采购说明"], [
        [
            r.get("timing_label"), r.get("suggested_order_date"), r.get("warehouse_name"), r.get("sku_no"), r.get("short_name"), _spec_name(r), _spec_remark(r), r.get("supplier_name"), r.get("production_line"), r.get("production_capacity"), r.get("production_days"), r.get("available_num"), r.get("purchase_in_transit_num"), r.get("transfer_qty"), r.get("trend_coefficient"), r.get("daily_sales"), r.get("calculated_order_qty"), r.get("moq"), r.get("moq_uplift_qty"), r.get("final_order_qty"),
            "暂不下单：转清仓评估" if r.get("low_demand_observation") else f"需运营确认：计算 {_number(r.get('calculated_order_qty')):.0f} < 起订 {_number(r.get('moq')):.0f}" if _number(r.get("moq_uplift_qty")) > 0 else "按计算量下单",
            _rounded_days(r.get("projected_coverage_days")), r.get("estimated_stockout_date"), r.get("plan_reason"),
        ] for r in data["purchase"]["items"]
    ], row_fill=lambda row: purchase_colors.get(str(purchase_severity.get((row[2], row[3]), ""))), integer_headers=["生产天数", "可用库存", "采购在途", "建议调拨", "计算采购量", "最低起订量", "起订量补足", "建议下单量", "剩余库存天数（含在途）"], decimal_headers=["趋势系数", "日销量"])
    # After adding 规格备注、供应商、剩余库存天数（含在途），起订量补足位于 S 列。
    moq_uplift_column = 19
    for row_index in range(2, purchase_sheet.max_row + 1):
        cell = purchase_sheet.cell(row_index, moq_uplift_column)
        if _number(cell.value) > 0:
            cell.fill = PatternFill("solid", fgColor=LIGHT_ORANGE)
    purchase_sheet.insert_rows(1, 4)
    purchase_sheet.merge_cells("A1:X1")
    purchase_sheet["A1"] = "🛒 采购计划总览"
    purchase_sheet["A1"].fill = PatternFill("solid", fgColor=GOLD)
    purchase_sheet["A1"].font = Font(color=WHITE, bold=True, size=16)
    purchase_sheet.merge_cells("A2:X2")
    purchase_sheet["A2"] = "包含五个运营仓全部启用的非清仓 SKU；红色=紧急排单，黄色=一周内，蓝色=一周外；缺生产周期不计算不着色，起订量补足单元格另行标色。"
    purchase_sheet["A2"].font = Font(color=GRAY, italic=True)
    summary_cells = [
        ("A3", "严重缺货", purchase_summary.get("severe_shortage_count", 0), RED),
        ("D3", "本周下单", purchase_summary.get("due_within_week_count", 0), ORANGE),
        ("G3", "计算采购量", purchase_summary.get("total_calculated_order_qty", 0), BLUE),
        ("J3", "起订量补足", purchase_summary.get("total_moq_uplift_qty", 0), GOLD),
        ("M3", "建议下单量", purchase_summary.get("total_order_qty", 0), GREEN),
        ("P3", "需运营确认", purchase_summary.get("moq_applied_count", 0), PURPLE),
    ]
    for anchor, label, amount, color in summary_cells:
        column = get_column_letter(purchase_sheet[anchor].column)
        start_row = purchase_sheet[anchor].row
        end_column = get_column_letter(purchase_sheet[anchor].column + 1)
        purchase_sheet.merge_cells(f"{column}{start_row}:{end_column}{start_row}")
        purchase_sheet[anchor] = label
        purchase_sheet[anchor].fill = PatternFill("solid", fgColor=color)
        purchase_sheet[anchor].font = Font(color=WHITE, bold=True)
        purchase_sheet[anchor].alignment = Alignment(horizontal="center")
        value = purchase_sheet.cell(start_row, purchase_sheet[anchor].column + 2, amount)
        value.font = Font(color=color, bold=True, size=12)
        value.number_format = "#,##0"
    purchase_sheet.auto_filter.ref = f"A5:X{purchase_sheet.max_row}"
    purchase_sheet.freeze_panes = "A6"
    purchase_sheet.row_dimensions[5].height = 34
    for cell in purchase_sheet[5]:
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    purchase_widths = {
        "A": 24, "B": 14, "C": 18, "D": 24, "E": 18, "F": 18, "G": 18, "H": 26,
        "I": 12, "J": 30, "K": 12, "L": 13, "M": 13, "N": 12,
        "O": 12, "P": 12, "Q": 14, "R": 14, "S": 14, "T": 14, "U": 30, "V": 16, "W": 14, "X": 42,
    }
    for column, width in purchase_widths.items():
        purchase_sheet.column_dimensions[column].width = width

    supplier: Dict[str, Dict[str, float]] = {}
    for row in data["purchase"]["items"]:
        key = row.get("supplier_name") or "供应商待补充"
        item = supplier.setdefault(key, {"sku": 0, "calculated": 0, "uplift": 0, "final": 0, "severe": 0})
        item["sku"] += 1
        item["calculated"] += _number(row.get("calculated_order_qty"))
        item["uplift"] += _number(row.get("moq_uplift_qty"))
        item["final"] += _number(row.get("final_order_qty"))
        item["severe"] += 1 if row.get("is_severe_shortage") else 0
    _write_table(workbook, "08_采购_供应商分析", ["供应商", "采购SKU", "严重缺货SKU", "计算采购量", "起订量补足", "建议下单量"], [
        [name, values["sku"], values["severe"], values["calculated"], values["uplift"], values["final"]]
        for name, values in sorted(supplier.items(), key=lambda item: (-item[1]["final"], item[0]))
    ], integer_headers=["采购SKU", "严重缺货SKU", "计算采购量", "起订量补足", "建议下单量"])

    _write_table(workbook, "09_采购_在途与入库", ["仓库", "SKU", "商品简称", "规格", "总入库", "采购入库", "调拨入库", "退货入库", "来源单数", "最后入库"], [
        [r.get("warehouse_name"), r.get("sku_no"), r.get("short_name"), _spec_name(r), r.get("inbound_qty"), r.get("purchase_qty"), r.get("transfer_qty"), r.get("return_qty"), r.get("order_count"), r.get("last_inbound_time")]
        for r in data["inbound"]["items"]
    ], integer_headers=["总入库", "采购入库", "调拨入库", "退货入库", "来源单数"])

    sync_values = database.last_sync() or {}
    sync_integer_fields = {"id", "movement_count", "inventory_count", "sales_count", "return_count", "cancellation_count"}
    _write_table(workbook, "10_共享_同步日志", ["字段", "值"], [[key, value] for key, value in sync_values.items()])
    sync_sheet = workbook.worksheets[-1]
    for row_index in range(2, sync_sheet.max_row + 1):
        if sync_sheet.cell(row_index, 1).value in sync_integer_fields:
            sync_sheet.cell(row_index, 2).value = _integer(sync_sheet.cell(row_index, 2).value)
            sync_sheet.cell(row_index, 2).number_format = "#,##0"
    _write_table(workbook, "11_共享_计算口径", ["模块", "指标", "口径"], [
        ["运营", "销量", "销售出库流水，按实际发货日期统计，保持仓库 + SKU 粒度。"],
        ["运营", "退货", "退货入库流水，按退货入库日期统计；退货仓仅用于退货统计。"],
        ["运营", "库存积压", "含采购在途的预计库存覆盖天数超过 90 天；按库存预警、中度积压、严重积压分级。"],
        ["运营", "结构调整", "仅生成运营候选，依据销量、趋势、库存覆盖、退货率和采购价/零售价毛利；不会自动修改四大结构。"],
        ["采购", "计算采购量", "按目标仓到货前缺口计算，已扣除可用库存、采购在途和建议调拨；达到最低起订量后按50件倍数取整，余数1-10向下、11及以上向上。"],
        ["采购", "起订量处理", "最终建议量取计算采购量与最低起订量的较大值；差额单独提示运营确认。"],
        ["采购", "采购计划", "五个运营仓全部启用 SKU；理论下单日按预计缺货日减完整交付周期倒推。含在途库存天数严格小于完整交付周期才标红；一周内黄色、一周外蓝色、清仓候选紫色；固定下单日仅作执行参考。"],
    ])
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def build_daily(database: Any, end_date: date) -> bytes:
    value = end_date.isoformat()
    return _build(database, value, value, "daily")


def build_weekly(database: Any, end_date: date) -> bytes:
    return _build(database, (end_date - timedelta(days=6)).isoformat(), end_date.isoformat(), "weekly")
