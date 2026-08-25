import tempfile
import unittest
from unittest.mock import patch
from datetime import date, datetime, timedelta
from io import BytesIO
from pathlib import Path
import sys

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))

from wangdian_inventory.app import create_app
from wangdian_inventory.config import Settings
from wangdian_inventory.db import InventoryDatabase, _is_lead_time_shortage, _round_purchase_qty_to_50
from wangdian_inventory.report_export import _rounded_days, _write_table
from wangdian_inventory.sync import InventorySynchronizer
from openpyxl import Workbook


class InventoryDatabaseTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temp_dir = tempfile.TemporaryDirectory()
        self.database = InventoryDatabase(Path(self.temp_dir.name) / "test.db")
        self.database.upsert_products(
            [
                {
                    "sku_no": "SKU-1",
                    "goods_name": "测试水杯",
                    "short_name": "水杯A",
                    "spec_name": "白色",
                    "retail_price": 100,
                },
                {
                    "sku_no": "SKU-2",
                    "goods_name": "测试雨伞",
                    "short_name": "雨伞A",
                    "spec_name": "蓝色",
                    "retail_price": 50,
                },
            ]
        )
        self.database.upsert_inventory(
            [
                {
                    "spec_no": "SKU-1",
                    "warehouse_id": "1",
                    "warehouse_name": "主仓",
                    "stock_num": 20,
                    "available_num": 18,
                    "avg_cost_price": 30,
                },
                {
                    "spec_no": "SKU-2",
                    "warehouse_id": "1",
                    "warehouse_name": "主仓",
                    "stock_num": 10,
                    "available_num": 10,
                    "avg_cost_price": 20,
                },
            ],
            "2026-07-30",
        )
        self.database.upsert_warehouses(
            [{"warehouse_id": "1", "warehouse_no": "MAIN", "name": "主仓", "is_disabled": 0}]
        )
        self.database.upsert_movements(
            [
                {
                    "sku_no": "SKU-1",
                    "warehouse_id": "1",
                    "in_out_type": "销售订单",
                    "out_num": 8,
                    "create_date": "2026-07-29 12:00:00",
                    "src_detail_id": "out-1",
                    "src_order_no": "JY1",
                },
                {
                    "sku_no": "SKU-1",
                    "warehouse_id": "1",
                    "in_out_type": "退货入库",
                    "in_num": 2,
                    "create_date": "2026-07-29 15:00:00",
                    "src_detail_id": "return-1",
                    "src_order_no": "RK1",
                },
                {
                    "sku_no": "SKU-2",
                    "warehouse_id": "1",
                    "in_out_type": "销售订单",
                    "out_num": 4,
                    "create_date": "2026-07-29 13:00:00",
                    "src_detail_id": "out-2",
                    "src_order_no": "JY2",
                },
                {
                    "sku_no": "SKU-1",
                    "warehouse_id": "1",
                    "warehouse_name": "主仓",
                    "in_out_type": "采购入库",
                    "in_num": 12,
                    "create_date": "2026-07-29 09:00:00",
                    "src_detail_id": "purchase-1",
                    "src_order_no": "CGRK1",
                },
                {
                    "sku_no": "SKU-2",
                    "warehouse_id": "1",
                    "warehouse_name": "主仓",
                    "in_out_type": "调拨入库",
                    "in_num": 3,
                    "create_date": "2026-07-29 10:00:00",
                    "src_detail_id": "transfer-1",
                    "src_order_no": "DBRK1",
                },
            ]
        )
        self.database.upsert_sales_orders(
            [
                {
                    "stockout_id": "SO-1",
                    "order_no": "CK-1",
                    "src_order_no": "JY1",
                    "shop_id": "10",
                    "shop_no": "P1",
                    "shop_name": "测试旗舰店",
                    "consign_time": "2026-07-29 12:00:00",
                    "warehouse_id": "1",
                    "status": 95,
                    "details_list": [
                        {
                            "rec_id": "out-1",
                            "src_order_detail_id": "out-1",
                            "spec_no": "SKU-1",
                            "num": 8,
                            "paid": 700,
                            "share_amount": 680,
                        }
                    ],
                },
                {
                    "stockout_id": "SO-2",
                    "order_no": "CK-2",
                    "src_order_no": "JY2",
                    "consign_time": "2026-07-29 13:00:00",
                    "warehouse_id": "1",
                    "status": 95,
                    "details_list": [
                        {
                            "rec_id": "out-2",
                            "src_order_detail_id": "out-2",
                            "spec_no": "SKU-2",
                            "num": 4,
                            "paid": 200,
                            "share_amount": 200,
                        }
                    ],
                },
            ]
        )
        self.database.upsert_return_orders(
            [
                {
                    "stockin_id": "SI-1",
                    "order_no": "RK-1",
                    "stockin_time": "2026-07-29 15:00:00",
                    "warehouse_id": "1",
                    "status": 80,
                    "details_list": [
                        {
                            "rec_id": "return-1",
                            "src_order_detail_id": "return-1",
                            "spec_no": "SKU-1",
                            "num": 2,
                            "actual_refund_amount": 150,
                        }
                    ],
                }
            ]
        )

    def tearDown(self) -> None:
        self.temp_dir.cleanup()

    def test_dashboard_calculates_net_sales_and_estimated_revenue(self) -> None:
        result = self.database.dashboard("2026-07-29", "2026-07-30")
        self.assertEqual(result["summary"]["stock_num"], 30)
        self.assertEqual(result["summary"]["sales_qty"], 12)
        self.assertEqual(result["summary"]["return_qty"], 2)
        self.assertEqual(result["summary"]["net_sales_qty"], 10)
        self.assertEqual(result["summary"]["estimated_revenue"], 800)
        self.assertEqual(result["summary"]["sales_amount"], 900)
        self.assertEqual(result["summary"]["refund_amount"], 150)
        self.assertEqual(result["summary"]["net_revenue"], 750)

    def test_sales_orders_persist_shop_fields(self) -> None:
        with self.database.connect() as connection:
            row = connection.execute(
                "SELECT shop_id, shop_no, shop_name FROM sales_lines WHERE stockout_id='SO-1'"
            ).fetchone()
        self.assertEqual(dict(row), {"shop_id": "10", "shop_no": "P1", "shop_name": "测试旗舰店"})

    def test_shop_sku_sales_groups_by_shop_and_sku(self) -> None:
        result = self.database.shop_sku_sales("2026-07-29", "2026-07-30", limit=20)
        self.assertEqual(result["summary"]["shop_count"], 2)
        self.assertEqual(result["summary"]["sku_count"], 2)
        p1 = next(row for row in result["items"] if row["shop_no"] == "P1")
        self.assertEqual(p1["sku_no"], "SKU-1")
        self.assertEqual(p1["sales_qty"], 8)
        self.assertEqual(p1["sales_amount"], 700)
        filtered = self.database.shop_sku_sales("2026-07-29", "2026-07-30", shop_no="P1", limit=20)
        self.assertEqual([row["shop_no"] for row in filtered["items"]], ["P1"])

    def test_search_filters_summary_chart_and_rows(self) -> None:
        result = self.database.dashboard(
            "2026-07-29", "2026-07-30", search="水杯"
        )
        self.assertEqual(len(result["items"]), 1)
        self.assertEqual(result["summary"]["sales_qty"], 8)
        self.assertEqual(result["daily"][0]["sales_qty"], 8)

    def test_inventory_status_pagination_and_warehouse_stats(self) -> None:
        self.database.upsert_inventory(
            [
                {
                    "spec_no": "SKU-2",
                    "warehouse_id": "1",
                    "warehouse_name": "主仓",
                    "stock_num": -2,
                    "available_num": -3,
                }
            ],
            "2026-07-30",
        )
        negative = self.database.dashboard(
            "2026-07-29", "2026-07-30", stock_status="negative"
        )
        self.assertEqual(negative["pagination"]["total"], 1)
        self.assertEqual(negative["items"][0]["sku_no"], "SKU-2")

        page = self.database.dashboard(
            "2026-07-29", "2026-07-30", limit=1, offset=1
        )
        self.assertEqual(page["pagination"]["total"], 2)
        self.assertEqual(len(page["items"]), 1)

        warehouse = self.database.warehouses()[0]
        self.assertEqual(warehouse["sku_count"], 2)

    def test_warehouse_sku_sales_uses_movement_quantities(self) -> None:
        result = self.database.warehouse_sku_sales(
            "2026-07-29", "2026-07-29", warehouse_id="1"
        )
        self.assertEqual(result["pagination"]["total"], 2)
        sku_1 = next(item for item in result["items"] if item["sku_no"] == "SKU-1")
        self.assertEqual(sku_1["warehouse_name"], "")
        self.assertEqual(sku_1["sales_qty"], 8)
        self.assertEqual(sku_1["return_qty"], 2)
        self.assertEqual(sku_1["net_sales_qty"], 6)
        self.assertEqual(sku_1["sales_7d_qty"], 8)
        self.assertEqual(sku_1["sales_15d_qty"], 8)
        self.assertEqual(sku_1["sales_30d_qty"], 8)
        self.assertEqual(result["summary"]["sales_7d_qty"], 12)
        self.assertEqual(result["summary"]["sales_15d_qty"], 12)
        self.assertEqual(result["summary"]["sales_30d_qty"], 12)

    def test_warehouse_sku_sales_adds_trend_and_transit_coverage(self) -> None:
        result = self.database.warehouse_sku_sales(
            "2026-07-29", "2026-07-29", warehouse_id="1"
        )
        sku_1 = next(item for item in result["items"] if item["sku_no"] == "SKU-1")
        # The fixture has 8 units in the 7/15/30-day rolling windows. The
        # coverage baseline is the 30-day average (8 / 30 per day), and the
        # current available stock is 18 with no transit in this fixture.
        self.assertAlmostEqual(sku_1["trend_coefficient"], 30 / 7, places=3)
        self.assertEqual(sku_1["inventory_with_transit_days"], 68)
        self.assertEqual(sku_1["estimated_stockout_date_with_transit"], "2026-10-04")

    def test_warehouse_sku_sales_includes_inventory_without_period_activity(self) -> None:
        self.database.upsert_products([{
            "sku_no": "SKU-INVENTORY-ONLY",
            "goods_name": "仅库存商品",
            "short_name": "仅库存",
            "spec_name": "红色",
        }])
        self.database.upsert_inventory([{
            "spec_no": "SKU-INVENTORY-ONLY",
            "warehouse_id": "1",
            "warehouse_name": "主仓",
            "stock_num": 9,
            "available_num": 9,
        }], "2026-07-30")
        result = self.database.warehouse_sku_sales(
            "2026-07-29", "2026-07-29", warehouse_id="1"
        )
        item = next(row for row in result["items"] if row["sku_no"] == "SKU-INVENTORY-ONLY")
        self.assertEqual(item["sales_qty"], 0)
        self.assertEqual(item["return_qty"], 0)
        self.assertEqual(item["stock_num"], 9)


    def test_inbound_analysis_separates_purchase_return_and_transfer(self) -> None:
        result = self.database.inbound_analysis(
            "2026-07-29", "2026-07-29", warehouse_id="1"
        )
        self.assertEqual(result["summary"]["inbound_qty"], 17)
        self.assertEqual(result["summary"]["purchase_qty"], 12)
        self.assertEqual(result["summary"]["return_qty"], 2)
        self.assertEqual(result["summary"]["transfer_qty"], 3)
        self.assertEqual(result["summary"]["sku_count"], 2)
        sku_1 = next(item for item in result["items"] if item["sku_no"] == "SKU-1")
        self.assertEqual(sku_1["inbound_qty"], 14)
        self.assertEqual(sku_1["purchase_qty"], 12)
        self.assertEqual(sku_1["return_qty"], 2)
        self.assertEqual(sku_1["warehouse_name"], "主仓")

        purchases = self.database.inbound_analysis(
            "2026-07-29", "2026-07-29", inbound_type=1
        )
        self.assertEqual(purchases["summary"]["inbound_qty"], 12)
        self.assertEqual(purchases["pagination"]["total"], 1)

    def test_returns_in_return_warehouse_are_visible_in_inbound_and_sales_views(self) -> None:
        self.database.upsert_warehouses([{
            "warehouse_id": "2", "warehouse_no": "RETURN", "name": "退货仓", "is_disabled": 0,
        }])
        self.database.upsert_movements([{
            "sku_no": "SKU-1", "warehouse_id": "2", "warehouse_name": "退货仓",
            "in_out_type": "退货入库", "in_num": 4,
            "create_date": "2026-07-29 18:00:00", "src_detail_id": "return-east-1",
            "src_order_no": "RK-RETURN-1",
        }])

        inbound = self.database.inbound_analysis("2026-07-29", "2026-07-29")
        self.assertEqual(inbound["summary"]["return_qty"], 6)
        return_row = next(row for row in inbound["items"] if row["warehouse_id"] == "2")
        self.assertEqual(return_row["return_qty"], 4)

        sales = self.database.warehouse_sku_sales("2026-07-29", "2026-07-29")
        return_row = next(row for row in sales["items"] if row["warehouse_id"] == "2")
        self.assertEqual(return_row["sales_qty"], 0)
        self.assertEqual(return_row["return_qty"], 4)
        self.assertEqual(return_row["net_sales_qty"], -4)

    def test_short_name_sales_aggregates_skus_and_warehouse_breakdown(self) -> None:
        self.database.upsert_products(
            [
                {
                    "sku_no": "SKU-2",
                    "goods_name": "测试雨伞",
                    "short_name": "水杯A",
                    "spec_name": "蓝色",
                }
            ]
        )
        result = self.database.short_name_sales("2026-07-29", "2026-07-29")
        grouped = next(item for item in result["items"] if item["display_name"] == "水杯A")
        self.assertEqual(result["pagination"]["total"], 1)
        self.assertEqual(grouped["sku_count"], 2)
        self.assertEqual(grouped["sales_qty"], 12)
        self.assertEqual(grouped["return_qty"], 2)
        self.assertEqual(grouped["net_sales_qty"], 10)
        self.assertEqual(grouped["stock_num"], 30)
        self.assertEqual(grouped["warehouses"][0]["sales_qty"], 12)

    def test_short_name_sales_marks_goods_number_fallback(self) -> None:
        self.database.upsert_products(
            [{"sku_no": "SKU-3", "goods_no": "GOODS-3", "goods_name": "测试帽子"}]
        )
        self.database.upsert_movements(
            [
                {
                    "sku_no": "SKU-3",
                    "warehouse_id": "1",
                    "in_out_type": "销售订单",
                    "out_num": 3,
                    "create_date": "2026-07-29 14:00:00",
                    "src_detail_id": "out-3",
                    "src_order_no": "JY3",
                }
            ]
        )
        result = self.database.short_name_sales(
            "2026-07-29", "2026-07-29", search="GOODS-3"
        )
        self.assertEqual(result["pagination"]["total"], 1)
        self.assertEqual(result["items"][0]["display_name"], "GOODS-3")
        self.assertEqual(result["items"][0]["is_fallback"], 1)

    def test_product_metadata_import_marks_unconfigured_skus(self) -> None:
        result = self.database.upsert_product_metadata(
            [
                {
                    "sku_no": "SKU-1",
                    "品类": "护具/护腕",
                    "四大结构": "基础款",
                    "起订量": "无",
                    "生产周期": "30天",
                },
                {
                    "sku_no": "NOT-IN-PRODUCTS",
                    "品类": "其他",
                    "四大结构": "测试款",
                    "起订量": "100",
                    "生产周期": "15天",
                },
            ],
            source="test.xlsx#8.5",
        )
        self.assertEqual(result, {"updated": 1, "matched": 1, "unmatched": 1})
        with self.database.connect() as connection:
            row = connection.execute(
                "SELECT category, product_structure, moq, production_days, metadata_status "
                "FROM products WHERE sku_no='SKU-1'"
            ).fetchone()
            other = connection.execute(
                "SELECT metadata_status FROM products WHERE sku_no='SKU-2'"
            ).fetchone()
        self.assertEqual(dict(row), {
            "category": "护具/护腕",
            "product_structure": "基础款",
            "moq": None,
            "production_days": 30,
            "metadata_status": "已配置",
        })
        self.assertEqual(other["metadata_status"], "待补充")

    def test_product_supplier_fields_are_saved_and_preserved(self) -> None:
        self.database.upsert_products([{
            "sku_no": "SKU-1",
            "provider_id": "19",
            "provider_no": "HB-19",
            "provider_name": "扬州市宏博体育用品有限公司",
        }])
        self.database.upsert_products([{"sku_no": "SKU-1", "goods_name": "测试水杯"}])
        with self.database.connect() as connection:
            row = connection.execute(
                "SELECT supplier_id, supplier_no, supplier_name, supplier_updated_at "
                "FROM products WHERE sku_no='SKU-1'"
            ).fetchone()
        self.assertEqual(dict(row)["supplier_id"], "19")
        self.assertEqual(dict(row)["supplier_no"], "HB-19")
        self.assertEqual(dict(row)["supplier_name"], "扬州市宏博体育用品有限公司")
        self.assertTrue(dict(row)["supplier_updated_at"])

    def test_supplier_backfill_updates_existing_skus_only(self) -> None:
        result = self.database.upsert_product_suppliers([
            {"spec_no": "SKU-1", "provider_name": "德清县新市镇铠博服装经营部"},
            {"spec_no": "MISSING", "provider_name": "未匹配供应商"},
        ])
        self.assertEqual(result["matched"], 1)
        self.assertEqual(result["with_supplier"], 2)
        with self.database.connect() as connection:
            row = connection.execute(
                "SELECT supplier_name FROM products WHERE sku_no='SKU-1'"
            ).fetchone()
        self.assertEqual(row["supplier_name"], "德清县新市镇铠博服装经营部")

    def test_replenishment_uses_enabled_sales_warehouse_inventory(self) -> None:
        self.database.upsert_inventory([{
            "spec_no": "SKU-1", "warehouse_id": "1", "warehouse_name": "主仓",
            "stock_num": 1000, "available_num": 1000,
        }], "2026-07-30")
        result = self.database.replenishment_analysis(
            "2026-07-29", "2026-07-29", limit=10
        )
        sku_1 = next(item for item in result["items"] if item["sku_no"] == "SKU-1")
        self.assertEqual(sku_1["available_num"], 1000)
        self.assertEqual(sku_1["sales_qty"], 8)
        self.assertEqual(sku_1["daily_sales"], 8)
        self.assertEqual(sku_1["coverage_days"], 125)
        self.assertEqual(sku_1["forecast_basis"], "近期平均")
        self.assertEqual(sku_1["recommendation"], "库存预警")

    def test_replenishment_is_grouped_by_sku_and_warehouse(self) -> None:
        self.database.upsert_warehouses([
            {"warehouse_id": "2", "warehouse_no": "EAST", "name": "东仓", "is_disabled": 0},
        ])
        self.database.upsert_inventory([
            {
                "spec_no": "SKU-1", "warehouse_id": "1", "warehouse_name": "主仓",
                "stock_num": 1000, "available_num": 1000,
            },
            {
                "spec_no": "SKU-1", "warehouse_id": "2", "warehouse_name": "东仓",
                "stock_num": 400, "available_num": 400,
            },
        ], "2026-07-30")
        self.database.upsert_movements([
            {
                "sku_no": "SKU-1", "warehouse_id": "2", "warehouse_name": "东仓",
                "in_out_type": "销售订单", "out_num": 3,
                "create_date": "2026-07-29 16:00:00", "src_detail_id": "east-out-1",
                "src_order_no": "JY-EAST-1",
            },
        ])

        result = self.database.replenishment_analysis(
            "2026-07-29", "2026-07-29", limit=20
        )
        sku_rows = [item for item in result["items"] if item["sku_no"] == "SKU-1"]
        self.assertEqual({item["warehouse_id"] for item in sku_rows}, {"1", "2"})
        by_warehouse = {item["warehouse_id"]: item for item in sku_rows}
        self.assertEqual(by_warehouse["1"]["available_num"], 1000)
        self.assertEqual(by_warehouse["2"]["available_num"], 400)
        self.assertEqual(by_warehouse["2"]["sales_qty"], 3)

        east_only = self.database.replenishment_analysis(
            "2026-07-29", "2026-07-29", warehouse_id="2", limit=20
        )
        self.assertEqual({item["warehouse_id"] for item in east_only["items"]}, {"2"})

    def test_inventory_alert_only_includes_skus_above_90_forecast_days_including_transit(self) -> None:
        self.database.upsert_products([
            {"sku_no": "SKU-90", "goods_name": "边界库存", "short_name": "边界", "spec_name": "90天"},
            {"sku_no": "SKU-91", "goods_name": "预警库存", "short_name": "预警", "spec_name": "91天"},
        ])
        self.database.upsert_inventory([
            {"spec_no": "SKU-90", "warehouse_id": "1", "warehouse_name": "主仓", "stock_num": 90, "available_num": 90},
            {"spec_no": "SKU-91", "warehouse_id": "1", "warehouse_name": "主仓", "stock_num": 91, "available_num": 91},
        ], "2026-07-30")
        self.database.upsert_movements([
            {"sku_no": "SKU-90", "warehouse_id": "1", "in_out_type": "销售订单", "out_num": 1, "create_date": "2026-07-29 10:00:00", "src_detail_id": "out-90", "src_order_no": "JY90"},
            {"sku_no": "SKU-91", "warehouse_id": "1", "in_out_type": "销售订单", "out_num": 1, "create_date": "2026-07-29 11:00:00", "src_detail_id": "out-91", "src_order_no": "JY91"},
        ])

        result = self.database.replenishment_analysis("2026-07-29", "2026-07-29", limit=20)
        sku_numbers = {item["sku_no"] for item in result["items"]}
        self.assertNotIn("SKU-90", sku_numbers)
        self.assertIn("SKU-91", sku_numbers)
        self.assertEqual(result["inventory_alert_threshold_days"], 90)
        self.assertEqual(result["pagination"]["total"], len(result["items"]))

        self.database.upsert_inventory([{
            "spec_no": "SKU-90", "warehouse_id": "1", "warehouse_name": "主仓",
            "stock_num": 90, "available_num": 90, "purchase_in_transit_num": 2,
        }], "2026-07-30")
        with_transit = self.database.replenishment_analysis("2026-07-29", "2026-07-29", limit=20)
        self.assertIn("SKU-90", {item["sku_no"] for item in with_transit["items"]})

    def test_replenishment_model_includes_transit_and_trend_fields(self) -> None:
        self.database.upsert_inventory(
            [{
                "spec_no": "SKU-1",
                "warehouse_id": "1",
                "warehouse_name": "主仓",
                "stock_num": 1000,
                "available_num": 1000,
                "purchase_in_transit_num": 12,
            }],
            "2026-07-30",
        )
        result = self.database.replenishment_analysis("2026-07-29", "2026-07-30", limit=10)
        item = next(row for row in result["items"] if row["sku_no"] == "SKU-1")
        self.assertEqual(item["sales_7d_qty"], 8)
        self.assertEqual(item["sales_15d_qty"], 8)
        self.assertEqual(item["sales_30d_qty"], 8)
        self.assertEqual(item["purchase_in_transit_num"], 12)
        self.assertAlmostEqual(item["trend_coefficient"], 30 / 7, places=3)
        self.assertEqual(item["forecast_basis"], "近期平均")
        self.assertAlmostEqual(item["forecast_daily_sales"], 8, places=3)
        self.assertGreater(item["projected_coverage_days"], item["forecast_coverage_days"])

    def test_clearance_products_are_excluded_from_replenishment(self) -> None:
        self.database.upsert_product_metadata(
            [{
                "sku_no": "SKU-1",
                "品类": "护具/护腕",
                "四大结构": "清仓款",
                "起订量": "无",
                "生产周期": "30天",
            }]
        )
        replenishment = self.database.replenishment_analysis("2026-07-29", "2026-07-30", limit=10)
        clearance = self.database.replenishment_analysis(
            "2026-07-29", "2026-07-30", mode="clearance", limit=10
        )
        self.assertNotIn("SKU-1", {item["sku_no"] for item in replenishment["items"]})
        item = next(item for item in clearance["items"] if item["sku_no"] == "SKU-1")
        self.assertEqual(item["suggested_restock"], 0)
        self.assertNotIn(item["recommendation"], {"缺货", "紧急补货", "建议补货"})

    def test_clearance_alert_requires_positive_low_stock_and_under_two_days_cover(self) -> None:
        self.database.upsert_product_metadata([{"sku_no": "SKU-1", "四大结构": "清仓款"}])
        self.database.upsert_inventory([{
            "spec_no": "SKU-1", "warehouse_id": "1", "warehouse_name": "主仓",
            "stock_num": 8, "available_num": 8,
        }], "2026-07-30")

        alert = self.database.replenishment_analysis("2026-07-29", "2026-07-29", limit=20)
        item = next(row for row in alert["items"] if row["sku_no"] == "SKU-1")
        self.assertEqual(item["recommendation"], "清仓预警")
        self.assertEqual(alert["summary"]["clearance_alert_count"], 1)

        filtered = self.database.replenishment_analysis(
            "2026-07-29", "2026-07-29", alert_status="清仓预警", limit=20
        )
        self.assertEqual([row["sku_no"] for row in filtered["items"]], ["SKU-1"])

        # 1-4 件即使库存天数尚未低于 2 天也应预警；零、负库存不预警。
        self.database.upsert_inventory([{
            "spec_no": "SKU-1", "warehouse_id": "1", "warehouse_name": "主仓",
            "stock_num": 4, "available_num": 4,
        }], "2026-07-30")
        low_quantity = self.database.replenishment_analysis("2026-07-29", "2026-07-29", limit=20)
        self.assertEqual(next(row for row in low_quantity["items"] if row["sku_no"] == "SKU-1")["recommendation"], "清仓预警")

        for available in (0, -1):
            self.database.upsert_inventory([{
                "spec_no": "SKU-1", "warehouse_id": "1", "warehouse_name": "主仓",
                "stock_num": available, "available_num": available,
            }], "2026-07-30")
            result = self.database.replenishment_analysis("2026-07-29", "2026-07-29", limit=20)
            self.assertNotIn("SKU-1", {row["sku_no"] for row in result["items"]})

    def test_clearance_summary_uses_saleable_warehouse_cost_and_weekly_snapshot(self) -> None:
        self.database.upsert_product_metadata([
            {"sku_no": "SKU-1", "四大结构": "清仓款"},
            {"sku_no": "SKU-2", "四大结构": "清仓款"},
        ])
        self.database.upsert_products([
            {"sku_no": "SKU-1", "sku_default_purchase_price": 30},
            {"sku_no": "SKU-2", "sku_default_purchase_price": 20},
        ])
        self.database.upsert_inventory([
            {
                "spec_no": "SKU-1", "warehouse_id": "1", "warehouse_name": "主仓",
                "stock_num": 10, "available_num": 8, "avg_cost_price": 30,
            },
            {
                "spec_no": "SKU-2", "warehouse_id": "1", "warehouse_name": "主仓",
                "stock_num": 5, "available_num": 5, "cost_price": 20,
            },
        ], "2026-08-03")
        self.database.upsert_warehouses([
            {"warehouse_id": "2", "warehouse_no": "RETURN", "name": "退货仓", "is_disabled": 0},
        ])
        self.database.upsert_inventory([
            {
                "spec_no": "SKU-1", "warehouse_id": "2", "warehouse_name": "退货仓",
                "stock_num": 99, "available_num": 99, "avg_cost_price": 30,
            },
        ], "2026-08-03")

        result = self.database.clearance_summary()
        self.assertEqual(result["summary"]["stock_num"], 15)
        self.assertEqual(result["summary"]["available_num"], 13)
        self.assertEqual(result["summary"]["purchase_cost"], 400)
        self.assertEqual(result["summary"]["missing_purchase_price_count"], 0)
        self.assertEqual(result["summary"]["sku_count"], 2)
        self.assertEqual(len(result["items"]), 2)

        self.assertEqual(self.database.record_clearance_weekly_snapshot("2026-08-03"), 2)
        updated = self.database.clearance_summary()
        self.assertEqual(updated["latest_weekly_snapshot"]["snapshot_date"], "2026-08-03")
        self.assertEqual(updated["latest_weekly_snapshot"]["purchase_cost"], 400)

    def test_goods_master_purchase_price_is_persisted_and_inventory_sync_does_not_clear_it(self) -> None:
        self.database.upsert_products([
            {"sku_no": "SKU-1", "sku_default_purchase_price": "12.5"},
        ])
        self.database.upsert_products([
            {"sku_no": "SKU-1", "goods_name": "测试水杯"},
        ])
        with self.database.connect() as connection:
            row = connection.execute(
                "SELECT purchase_price FROM products WHERE sku_no='SKU-1'"
            ).fetchone()
        self.assertEqual(row["purchase_price"], 12.5)

    def test_clearance_lists_unfinished_items_first_and_filters_by_status(self) -> None:
        self.database.upsert_products([
            {"sku_no": "SKU-3", "goods_name": "测试靠垫", "short_name": "靠垫A", "spec_name": "灰色"},
        ])
        self.database.upsert_inventory([
            {"spec_no": "SKU-2", "warehouse_id": "1", "warehouse_name": "主仓", "stock_num": 0, "available_num": 0},
            {"spec_no": "SKU-3", "warehouse_id": "1", "warehouse_name": "主仓", "stock_num": 5, "available_num": 5},
        ], "2026-07-30")
        self.database.upsert_product_metadata([
            {"sku_no": "SKU-1", "四大结构": "清仓款"},
            {"sku_no": "SKU-2", "四大结构": "清仓款"},
            {"sku_no": "SKU-3", "四大结构": "清仓款"},
        ])

        result = self.database.replenishment_analysis(
            "2026-07-29", "2026-07-30", mode="clearance", limit=10
        )
        self.assertEqual(
            [item["recommendation"] for item in result["items"]],
            ["清仓进行中", "清仓停滞", "已清完"],
        )

        stagnant = self.database.replenishment_analysis(
            "2026-07-29", "2026-07-30", mode="clearance", clearance_status="stagnant", limit=10
        )
        self.assertEqual([item["sku_no"] for item in stagnant["items"]], ["SKU-3"])

    def test_warehouse_plan_transfers_before_direct_purchase_and_excludes_clearance(self) -> None:
        self.database.upsert_warehouses([
            {
                "warehouse_id": "1", "warehouse_no": "MAIN", "name": "主仓库（新）",
                "is_disabled": 0,
            },
            {
                "warehouse_id": "2", "warehouse_no": "JP", "name": "日本仓",
                "is_disabled": 0, "transfer_source_enabled": 1,
            },
        ])
        self.database.upsert_inventory([
            {"spec_no": "SKU-1", "warehouse_id": "1", "warehouse_name": "主仓", "available_num": 1},
            {"spec_no": "SKU-1", "warehouse_id": "2", "warehouse_name": "东仓", "available_num": 20},
        ], "2026-07-30")
        self.database.upsert_product_metadata([{
            "sku_no": "SKU-1", "四大结构": "其他款", "生产周期": "30天",
        }])
        plan = self.database.purchase_plan("2026-07-29", "2026-07-29", limit=20)
        target = next(row for row in plan["items"] if row["sku_no"] == "SKU-1" and row["warehouse_id"] == "1")
        self.assertGreater(target["transfer_qty"], 0)
        self.assertGreater(target["final_order_qty"], 0)
        transfer = self.database.transfer_plan("2026-07-29", "2026-07-29", limit=20)
        self.assertEqual(transfer["items"][0]["source_warehouse_id"], "2")
        self.assertEqual(transfer["items"][0]["target_warehouse_id"], "1")

        self.database.upsert_product_metadata([{"sku_no": "SKU-1", "四大结构": "清仓款"}])
        clearance = self.database.purchase_plan("2026-07-29", "2026-07-29", limit=20)
        self.assertNotIn("SKU-1", {row["sku_no"] for row in clearance["items"]})

    def test_purchase_plan_exposes_calculated_quantity_before_moq_uplift(self) -> None:
        self.database.upsert_warehouses([{
            "warehouse_id": "1", "warehouse_no": "MAIN", "name": "主仓库（新）", "is_disabled": 0,
        }])
        self.database.upsert_product_metadata([{
            "sku_no": "SKU-1", "四大结构": "其他款", "生产周期": "30天", "起订量": 500,
        }])
        plan = self.database.purchase_plan("2026-07-29", "2026-07-29", limit=20)
        item = next(row for row in plan["items"] if row["sku_no"] == "SKU-1")
        self.assertGreater(item["calculated_order_qty"], 0)
        self.assertLess(item["calculated_order_qty"], 500)
        self.assertEqual(item["final_order_qty"], 500)
        self.assertEqual(item["moq_uplift_qty"], 500 - item["calculated_order_qty"])
        self.assertTrue(item["moq_applied"])

    def test_hongbo_defaults_to_fixed_thirty_day_cycle_and_uses_configured_moq(self) -> None:
        self.database.upsert_warehouses([{
            "warehouse_id": "1", "warehouse_no": "MAIN", "name": "主仓库（新）", "is_disabled": 0,
        }])
        self.database.upsert_product_metadata([{
            "sku_no": "SKU-1", "四大结构": "基础款", "起订量": 100,
        }])
        self.database.upsert_product_suppliers([{
            "sku_no": "SKU-1",
            "provider_name": "扬州市宏博体育用品有限公司",
        }])
        plan = self.database.purchase_plan("2026-07-29", "2026-07-29", limit=20)
        item = next(row for row in plan["items"] if row["sku_no"] == "SKU-1")
        self.assertEqual(item["production_days"], 30)
        self.assertEqual(item["moq"], 100)
        self.assertEqual(item["order_window"], "固定下单日（每月5日、15日）")
        self.assertEqual(item["lead_days"], 40)

    def test_purchase_quantity_rounding_uses_ten_unit_cutoff(self) -> None:
        self.assertEqual(_round_purchase_qty_to_50(103), 100)
        self.assertEqual(_round_purchase_qty_to_50(109), 100)
        self.assertEqual(_round_purchase_qty_to_50(111), 150)
        self.assertEqual(_round_purchase_qty_to_50(238), 250)

    def test_severe_shortage_uses_strict_full_lead_time_boundary(self) -> None:
        self.assertFalse(_is_lead_time_shortage(50, 40))
        self.assertFalse(_is_lead_time_shortage(40, 40))
        self.assertTrue(_is_lead_time_shortage(39.99, 40))

    def test_report_integer_columns_and_coverage_days_are_whole_numbers(self) -> None:
        self.assertEqual(_rounded_days(90.49), 90)
        self.assertEqual(_rounded_days(90.5), 91)

        workbook = Workbook()
        sheet = _write_table(
            workbook,
            "格式校验",
            ["可用库存", "预计剩余天数", "趋势系数"],
            [[18.0, 90.5, 1.234]],
            integer_headers=["可用库存", "预计剩余天数"],
            decimal_headers=["趋势系数"],
        )
        self.assertEqual([sheet["A2"].value, sheet["B2"].value], [18, 91])
        self.assertEqual([sheet["A2"].number_format, sheet["B2"].number_format], ["#,##0", "#,##0"])
        self.assertEqual(sheet["C2"].number_format, "#,##0.###")

    def test_purchase_plan_low_demand_candidate_is_not_reordered_as_shortage_purchase(self) -> None:
        self.database.upsert_warehouses([{
            "warehouse_id": "1", "warehouse_no": "MAIN", "name": "主仓库（新）", "is_disabled": 0,
        }])
        self.database.upsert_product_metadata([{
            "sku_no": "SKU-1", "四大结构": "其他款", "生产周期": "30天", "起订量": 100,
        }])
        self.database.upsert_inventory([{
            "spec_no": "SKU-1", "warehouse_id": "1", "warehouse_name": "主仓",
            "stock_num": 0, "available_num": 0,
        }], "2026-08-16")
        # Sales history exists across the business, while this SKU itself has
        # no shipment in the latest seven days.
        self.database.upsert_movements([{
            "sku_no": "SKU-2", "warehouse_id": "1", "warehouse_name": "主仓",
            "in_out_type": "销售订单", "out_num": 1,
            "create_date": f"{(date(2026, 8, 16) - timedelta(days=offset)).isoformat()} 12:00:00",
            "src_detail_id": f"history-{offset}", "src_order_no": f"HISTORY-{offset}",
        } for offset in range(30)])

        plan = self.database.purchase_plan("2026-08-16", "2026-08-16", limit=20)
        item = next(row for row in plan["items"] if row["sku_no"] == "SKU-1")
        self.assertTrue(item["low_demand_observation"])
        self.assertEqual(item["timing_label"], "低销量观察（清仓候选）")
        self.assertEqual(item["final_order_qty"], 0)
        self.assertIsNone(item["suggested_order_date"])

    def test_disabled_and_return_warehouses_do_not_enter_operational_stock_or_transfers(self) -> None:
        self.database.upsert_warehouses([
            {"warehouse_id": "2", "warehouse_no": "DISABLED", "name": "京东POP仓", "is_disabled": 0},
            {"warehouse_id": "3", "warehouse_no": "RETURN", "name": "退货仓", "is_disabled": 0},
        ])
        self.database.upsert_inventory([
            {"spec_no": "SKU-1", "warehouse_id": "2", "warehouse_name": "京东POP仓", "stock_num": 999, "available_num": 999},
            {"spec_no": "SKU-1", "warehouse_id": "3", "warehouse_name": "退货仓", "stock_num": 888, "available_num": 888},
        ], "2026-07-30")
        with self.database.connect() as connection:
            rows = connection.execute(
                "SELECT warehouse_name,is_disabled,transfer_source_enabled FROM warehouse_master WHERE warehouse_id IN ('2','3')"
            ).fetchall()
        self.assertEqual({row["warehouse_name"]: (row["is_disabled"], row["transfer_source_enabled"]) for row in rows}, {
            "京东POP仓": (1, 0), "退货仓": (0, 0),
        })
        dashboard = self.database.dashboard("2026-07-29", "2026-07-30")
        sku = next(row for row in dashboard["items"] if row["sku_no"] == "SKU-1")
        self.assertEqual(sku["available_num"], 18)

    def test_dashboard_includes_returns_received_by_active_return_warehouse(self) -> None:
        self.database.upsert_warehouses([
            {"warehouse_id": "3", "warehouse_no": "RETURN", "name": "退货仓", "is_disabled": 0},
        ])
        self.database.upsert_movements([
            {
                "sku_no": "SKU-1",
                "warehouse_id": "3",
                "warehouse_name": "退货仓",
                "in_out_type": "退货入库",
                "in_num": 3,
                "create_date": "2026-07-29 16:00:00",
                "src_detail_id": "return-warehouse-1",
                "src_order_no": "RK-RETURN-1",
            },
        ])

        dashboard = self.database.dashboard("2026-07-29", "2026-07-29")
        self.assertEqual(dashboard["summary"]["movement_return_qty"], 5)
        self.assertEqual(dashboard["daily"][0]["movement_return_qty"], 5)

    def test_sku_detail_uses_sales_and_return_facts(self) -> None:
        result = self.database.sku_detail("SKU-1", "2026-07-29", "2026-07-29")
        self.assertIsNotNone(result)
        self.assertEqual(result["financials"]["sales_amount"], 700)
        self.assertEqual(result["financials"]["refund_amount"], 150)
        self.assertEqual(result["daily"][0]["sales_qty"], 8)
        self.assertEqual(result["daily"][0]["return_qty"], 2)

    def test_upserting_same_movement_is_idempotent(self) -> None:
        movement = {
            "sku_no": "SKU-1",
            "warehouse_id": "1",
            "in_out_type": "销售订单",
            "out_num": 5,
            "create_date": "2026-07-30 10:00:00",
            "src_detail_id": "same-detail",
            "src_order_no": "JY3",
        }
        self.database.upsert_movements([movement])
        self.database.upsert_movements([movement])
        result = self.database.dashboard("2026-07-30", "2026-07-30")
        self.assertEqual(result["summary"]["movement_sales_qty"], 5)
        with self.database.connect() as connection:
            count = connection.execute(
                "SELECT COUNT(*) FROM movements WHERE source_detail_id='same-detail'"
            ).fetchone()[0]
        self.assertEqual(count, 1)

    def test_split_movement_rows_in_different_warehouses_are_preserved(self) -> None:
        base = {
            "sku_no": "SKU-1",
            "in_out_type": "销售订单",
            "out_num": 2,
            "create_date": "2026-07-30 10:00:00",
            "src_detail_id": "split-detail",
            "src_order_no": "JY-SPLIT",
        }
        self.database.upsert_movements(
            [
                {**base, "warehouse_id": "1", "position_id": "P1"},
                {**base, "warehouse_id": "2", "position_id": "P2"},
            ]
        )
        with self.database.connect() as connection:
            count = connection.execute(
                "SELECT COUNT(*) FROM movements WHERE source_detail_id='split-detail'"
            ).fetchone()[0]
        self.assertEqual(count, 2)

    def test_sales_reversal_is_kept_as_movement_audit(self) -> None:
        self.database.upsert_movements(
            [
                {
                    "sku_no": "SKU-1",
                    "warehouse_id": "1",
                    "in_out_type": "销售订单",
                    "src_order_type": 1,
                    "out_num": -3,
                    "create_date": "2026-07-29 18:00:00",
                    "src_detail_id": "out-1",
                    "src_order_no": "JY1",
                }
            ]
        )
        result = self.database.dashboard("2026-07-29", "2026-07-29")
        self.assertEqual(result["summary"]["sales_qty"], 12)
        self.assertEqual(result["summary"]["sales_amount"], 900)
        self.assertEqual(result["summary"]["movement_sales_qty"], 9)

    def test_cancelled_sales_line_is_excluded(self) -> None:
        self.database.upsert_sales_orders(
            [
                {
                    "stockout_id": "SO-1",
                    "order_no": "CK-1",
                    "consign_time": "2026-07-29 12:00:00",
                    "warehouse_id": "1",
                    "status": 5,
                    "details_list": [
                        {
                            "rec_id": "out-1",
                            "src_order_detail_id": "out-1",
                            "spec_no": "SKU-1",
                            "num": 8,
                            "paid": 700,
                        }
                    ],
                }
            ]
        )
        result = self.database.dashboard("2026-07-29", "2026-07-29")
        self.assertEqual(result["summary"]["sales_qty"], 4)
        self.assertEqual(result["summary"]["sales_amount"], 200)

    def test_replacing_inventory_removes_stale_rows(self) -> None:
        self.database.replace_inventory(
            [
                {
                    "spec_no": "SKU-1",
                    "warehouse_id": "1",
                    "stock_num": 7,
                    "available_num": 6,
                }
            ],
            "2026-07-31",
        )
        result = self.database.dashboard("2026-07-29", "2026-07-31")
        self.assertEqual(result["summary"]["stock_num"], 7)
        sku_2 = next(item for item in result["items"] if item["sku_no"] == "SKU-2")
        self.assertEqual(sku_2["stock_num"], 0)

    def test_inventory_combines_good_and_defective_stock(self) -> None:
        self.database.replace_inventory(
            [
                {"spec_no": "SKU-1", "warehouse_id": "1", "stock_num": 19, "available_num": 19, "defect": 0},
                {"spec_no": "SKU-1", "warehouse_id": "1", "stock_num": 100, "available_num": 0, "defect": 1},
            ],
            "2026-07-31",
        )
        result = self.database.dashboard("2026-07-29", "2026-07-31")
        sku_1 = next(item for item in result["items"] if item["sku_no"] == "SKU-1")
        self.assertEqual(sku_1["stock_num"], 119)
        self.assertEqual(sku_1["available_num"], 19)

    def test_inventory_ignores_impossible_sentinel_values(self) -> None:
        self.database.replace_inventory(
            [
                {
                    "spec_no": "SKU-1",
                    "warehouse_id": "1",
                    "stock_num": -100000005010,
                    "available_num": -100000005035,
                }
            ],
            "2026-07-31",
        )
        result = self.database.dashboard("2026-07-29", "2026-07-31")
        sku_1 = next(item for item in result["items"] if item["sku_no"] == "SKU-1")
        self.assertEqual(sku_1["stock_num"], 0)
        self.assertEqual(sku_1["available_num"], 0)


class InventoryApiTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temp_dir = tempfile.TemporaryDirectory()
        settings = Settings(
            sid="",
            app_key="",
            app_secret="",
            environment="test",
            database_path=Path(self.temp_dir.name) / "demo.db",
            demo_mode=True,
        )
        self.application = create_app(settings)

    def get_json(self, path: str) -> tuple[dict, int]:
        from urllib.parse import parse_qs, urlparse
        parsed = urlparse(path)
        query = {key: values[-1] for key, values in parse_qs(parsed.query).items() if values}
        if parsed.path == "/api/status":
            return self.application.json_status(), 200
        if parsed.path == "/api/dashboard":
            return self.application.dashboard(query)
        if parsed.path == "/api/warehouse-sales":
            return self.application.warehouse_sales(query)
        if parsed.path == "/api/shop-sales":
            return self.application.shop_sales(query)
        if parsed.path == "/api/inbound":
            return self.application.inbound(query)
        if parsed.path == "/api/short-name-sales":
            return self.application.short_name_sales(query)
        if parsed.path == "/api/replenishment":
            return self.application.replenishment(query)
        if parsed.path == "/api/purchase-plan":
            return self.application.purchase_plan(query)
        if parsed.path == "/api/transfer-plan":
            return self.application.transfer_plan(query)
        if parsed.path == "/api/clearance":
            return self.application.clearance(query)
        if parsed.path == "/api/clearance-summary":
            return self.application.clearance_summary(query)
        if parsed.path == "/api/export.csv":
            body, _, status = self.application.export_csv(query)
            return {"body": body}, status
        raise AssertionError(path)

    def tearDown(self) -> None:
        self.temp_dir.cleanup()

    def test_dashboard_and_status_endpoints(self) -> None:
        status, status_code = self.get_json("/api/status")
        self.assertEqual(status_code, 200)
        self.assertTrue(status["demo_mode"])

        dashboard, dashboard_code = self.get_json(
            f"/api/dashboard?start={date.today().isoformat()}&end={date.today().isoformat()}"
        )
        self.assertEqual(dashboard_code, 200)
        self.assertGreater(len(dashboard["items"]), 0)

    def test_daily_and_weekly_report_exports_are_xlsx_files(self) -> None:
        for report_type in ("daily", "weekly"):
            content, status, filename = self.application.export_report(
                report_type, {"date": date.today().isoformat()}
            )
            self.assertEqual(status, 200)
            self.assertTrue(filename.endswith(".xlsx"))
            self.assertTrue(content.startswith(b"PK"))

        content, status, _ = self.application.export_report(
            "daily", {"date": date.today().isoformat()}
        )
        self.assertEqual(status, 200)
        sales_sheet = load_workbook(BytesIO(content))["04_运营_采购在途预警"]
        headers = [cell.value for cell in sales_sheet[1]]
        self.assertEqual(
            headers,
            ["仓库", "SKU", "商品简称", "规格", "规格备注", "供应商", "日销量", "净销量", "退货率", "7日销量", "15日销量", "30日销量", "库存", "可用库存", "采购在途", "不含在途库存天数", "含在途库存天数"],
        )
        self.assertNotIn("退货量", headers)
        for sheet_name in (
            "02_运营_清仓预警", "03_运营_库存积压预警", "04_运营_采购在途预警",
            "05_运营_结构调整建议", "06_运营_零销量与新品", "07_采购_采购计划",
            "09_采购_在途与入库",
        ):
            sheet = load_workbook(BytesIO(content))[sheet_name]
            header_row = 5 if sheet_name == "07_采购_采购计划" else 1
            self.assertIn("规格", [cell.value for cell in sheet[header_row]])
        allowed_warehouses = {"主仓库（新）", "TMT康复仓", "杰菲克德国仓", "日本仓", "ENYISA仓"}
        self.assertTrue(
            all(sales_sheet.cell(row, 1).value in allowed_warehouses for row in range(2, sales_sheet.max_row + 1))
        )
        inventory_days_column = headers.index("不含在途库存天数") + 1
        self.assertTrue(
            all(
                sales_sheet.cell(row, inventory_days_column).number_format == "#,##0"
                for row in range(2, sales_sheet.max_row + 1)
                if sales_sheet.cell(row, inventory_days_column).value is not None
            )
        )
        transit_column = headers.index("采购在途") + 1
        self.assertTrue(
            all(
                sales_sheet.cell(row, transit_column).fill.fgColor.rgb in {"00FFF2CC", "FFF2CC"}
                for row in range(2, sales_sheet.max_row + 1)
                if sales_sheet.cell(row, transit_column).value and sales_sheet.cell(row, transit_column).value > 0
            )
        )

        sales, sales_code = self.get_json(
            f"/api/warehouse-sales?start={date.today().isoformat()}&end={date.today().isoformat()}"
        )
        self.assertEqual(sales_code, 200)
        self.assertIn("pagination", sales)

        shop_sales, shop_sales_code = self.get_json(
            f"/api/shop-sales?start={date.today().isoformat()}&end={date.today().isoformat()}"
        )
        self.assertEqual(shop_sales_code, 200)
        self.assertIn("shops", shop_sales)

        inbound, inbound_code = self.get_json(
            f"/api/inbound?start={date.today().isoformat()}&end={date.today().isoformat()}"
        )
        self.assertEqual(inbound_code, 200)
        self.assertIn("daily", inbound)

        short_names, short_names_code = self.get_json(
            f"/api/short-name-sales?start={date.today().isoformat()}&end={date.today().isoformat()}"
        )
        self.assertEqual(short_names_code, 200)
        self.assertIn("pagination", short_names)

        replenishment, replenishment_code = self.get_json("/api/replenishment")
        self.assertEqual(replenishment_code, 200)
        self.assertIn("target_days", replenishment)

        clearance, clearance_code = self.get_json("/api/clearance")
        self.assertEqual(clearance_code, 200)
        self.assertEqual(clearance["mode"], "clearance")

        clearance_summary, clearance_summary_code = self.get_json("/api/clearance-summary")
        self.assertEqual(clearance_summary_code, 200)
        self.assertIn("purchase_cost", clearance_summary["summary"])

    def test_inbound_rejects_invalid_type(self) -> None:
        response, status_code = self.get_json("/api/inbound?inbound_type=-1")
        self.assertEqual(status_code, 400)
        self.assertIn("error", response)

    def test_sync_requires_credentials(self) -> None:
        response, status_code = self.application.sync({"date": date.today().isoformat()})
        self.assertEqual(status_code, 409)

    def test_inventory_only_sync_scope(self) -> None:
        self.application.settings = Settings(
            sid="sid", app_key="key", app_secret="secret",
            environment="test", database_path=self.application.settings.database_path,
            demo_mode=False,
        )
        with patch("wangdian_inventory.app.InventorySynchronizer") as synchronizer:
            synchronizer.return_value.sync_inventory_only.return_value = {
                "status": "success", "scope": "inventory", "inventory_count": 2,
            }
            response, status_code = self.application.sync({"scope": "inventory"})
        self.assertEqual(status_code, 200)
        self.assertEqual(response["scope"], "inventory")

    def test_csv_export_has_utf8_bom(self) -> None:
        response, status_code = self.get_json(
            f"/api/export.csv?start={date.today().isoformat()}&end={date.today().isoformat()}"
        )
        self.assertEqual(status_code, 200)
        self.assertTrue(response["body"].startswith("\ufeff"))


class InventorySynchronizerTests(unittest.TestCase):
    def test_api_duplicate_movement_rows_are_counted_once_per_response(self) -> None:
        temp_dir = tempfile.TemporaryDirectory()
        self.addCleanup(temp_dir.cleanup)
        database = InventoryDatabase(Path(temp_dir.name) / "test.db")
        duplicate = {
            "sku_no": "SKU-DUP",
            "warehouse_id": "39",
            "warehouse_name": "主仓库（新）",
            "in_out_type": "销售订单",
            "src_id": "ORDER-ID",
            "src_detail_id": "DETAIL-ID",
            "src_order_no": "JY-DUP",
            "stockout_no": "JY-DUP",
            "out_num": "8",
            "num": "8",
            "create_date": "2026-07-30 14:19:49",
            "position_id": "POSITION-ID",
            "position_no": "A01-01-01-01",
        }

        class ClientStub:
            def call(self, endpoint, parameters):
                self.endpoint = endpoint
                return {"code": 0, "total_count": 2, "data": [duplicate, dict(duplicate)]}

        settings = Settings(
            sid="sid", app_key="key", app_secret="secret",
            environment="test", database_path=Path(temp_dir.name) / "test.db",
            demo_mode=False,
        )
        synchronizer = InventorySynchronizer(settings, database)
        client = ClientStub()

        synchronizer.sync_movements(client, date(2026, 7, 30))
        synchronizer.sync_movements(client, date(2026, 7, 30))

        with database.connect() as connection:
            count = connection.execute(
                "SELECT COUNT(*) FROM movements WHERE src_order_no='JY-DUP'"
            ).fetchone()[0]
            quantity = connection.execute(
                "SELECT SUM(out_num) FROM movements WHERE src_order_no='JY-DUP'"
            ).fetchone()[0]
        self.assertEqual(count, 2)
        self.assertEqual(quantity, 16)

    def test_goods_sync_uses_goods_short_name_when_spec_value_is_empty(self) -> None:
        class DatabaseStub:
            def product_skus(self):
                return ["SKU-1"]

            def upsert_products(self, records):
                self.records = list(records)
                return len(self.records)

        class ClientStub:
            def call(self, endpoint, parameters):
                return {
                    "code": 0,
                    "total_count": 1,
                    "goods_list": [
                        {
                            "goods_no": "GOODS-1",
                            "goods_name": "测试商品",
                            "short_name": "简称A",
                            "spec_list": [
                                {"spec_no": "SKU-1", "short_name": ""}
                            ],
                        }
                    ],
                }

        settings = Settings(
            sid="sid", app_key="key", app_secret="secret",
            environment="test", database_path=Path("unused.db"), demo_mode=False,
        )
        database = DatabaseStub()
        count = InventorySynchronizer(settings, database).sync_goods(
            ClientStub(), full=True
        )
        self.assertEqual(count, 1)
        self.assertEqual(database.records[0]["short_name"], "简称A")

    def test_inventory_sends_documented_json_string_for_sku_list(self) -> None:
        class DatabaseStub:
            def product_skus(self):
                return ["普通SKU"]

            def replace_inventory(self, records, snapshot_date):
                self.records = records
                return len(records)

        class ClientStub:
            def call(self, endpoint, parameters):
                self.parameters = parameters
                return {"code": 0, "stocks": []}

        settings = Settings(
            sid="sid", app_key="key", app_secret="secret",
            environment="test", database_path=Path("unused.db"), demo_mode=False,
        )
        database = DatabaseStub()
        client = ClientStub()
        InventorySynchronizer(settings, database).sync_inventory(client)
        self.assertEqual(client.parameters["spec_no_list"], r'["\u666e\u901aSKU"]')

    def test_empty_inventory_uses_documented_time_range(self) -> None:
        class DatabaseStub:
            def product_skus(self):
                return []

            def replace_inventory(self, records, snapshot_date):
                self.records = records
                return len(records)

        class ClientStub:
            def call(self, endpoint, parameters):
                self.calls.append((endpoint, parameters))
                return {"code": 0, "stocks": []}

            calls = []

        settings = Settings(
            sid="sid", app_key="key", app_secret="secret",
            environment="production", database_path=Path("unused.db"), demo_mode=False,
        )
        database = DatabaseStub()
        client = ClientStub()
        synchronizer = InventorySynchronizer(settings, database)
        synchronizer.INITIAL_INVENTORY_START = datetime.now().replace(microsecond=0)
        synchronizer.sync_inventory(client)
        endpoint, parameters = client.calls[0]
        self.assertEqual(endpoint, "stock_query")
        self.assertIn("start_time", parameters)
        self.assertIn("end_time", parameters)
        self.assertNotIn("spec_no_list", parameters)

    def test_full_inventory_ignores_partial_product_list(self) -> None:
        class DatabaseStub:
            def product_skus(self):
                return ["PARTIAL-SKU"]

            def replace_inventory(self, records, snapshot_date):
                return len(records)

        class ClientStub:
            def __init__(self):
                self.calls = []

            def call(self, endpoint, parameters):
                self.calls.append((endpoint, parameters))
                return {"code": 0, "stocks": []}

        settings = Settings(
            sid="sid", app_key="key", app_secret="secret",
            environment="production", database_path=Path("unused.db"), demo_mode=False,
        )
        client = ClientStub()
        synchronizer = InventorySynchronizer(settings, DatabaseStub())
        synchronizer.INITIAL_INVENTORY_START = datetime.now().replace(microsecond=0)
        synchronizer.sync_inventory(client, full=True)
        self.assertIn("start_time", client.calls[0][1])
        self.assertNotIn("spec_no_list", client.calls[0][1])


if __name__ == "__main__":
    unittest.main()
